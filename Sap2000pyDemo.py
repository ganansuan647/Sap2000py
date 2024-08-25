from pathlib import Path

import numpy as np
import openpyxl
from rich import print

from Sap2000py import Saproject

#full path to the model
ModelPath = Path('.\Test\Test.sdb')

# Create a Sap2000py obj (default: attatch to instance and create if not exist)
Sap = Saproject()

# Change Sap api settings if you want
# Sap.createSap(AttachToInstance = True,SpecifyPath = False,ProgramPath = "your sap2000 path")

# Open Sap2000program
Sap.openSap()

# Open Sap2000 sdb file (create if not exist, default: CurrentPath\\NewSapProj.sdb)
Sap.File.Open(ModelPath)
# And print the project information
Sap.getProjectInfo()

# Can also creat a new blank model (or other models you need like 2DFrame/3DFrame etc.)
Sap.File.New_Blank()


# Get the model information (getSapVersion() will print and also returns the version)
vsesion = Sap.SapVersion
Sap.getSapVersion()

# Check your units (getUnits() will print and also returns the units)
unit = Sap.Units
Sap.getUnits()
# set Units as you wish (Just type in Sap.setUnits(""), it will show you the options in your IDE)
Sap.setUnits("KN_m_C")

# Add China Common Material Set·
Sap.Scripts.AddCommonMaterialSet(standard = "JTG")


# Build your Model Here
# Add Joints by Script
joint_coord = np.array([[0,0,0],[10,0,0],[20,0,0],[30,0,0]])
Sap.Scripts.AddJoints(joint_coord)
# You can also Add Joints once a time : Sap.Assign.PointObj.AddCartesian(x=0,y=0,z=0)
# After using this script to add joints, you can see all the joints in var Sap.coord_joints

# Build Elements by Script
Sap.Scripts.AddElements([[1,2],[2,3],[3,4]])


# Add elements to your group
Sap.Scripts.Group.AddtoGroup('Edge',['1','4'],"Point")
# Check Your Group Elements
Eledict = Sap.Scripts.Group.GetElements('Edge')
print(Eledict)
# Select the group you need
Sap.Scripts.Group.Select('Edge')


# Modal Analysis
# Remove all cases for analysis
Sap.Scripts.Analyze.RemoveCases("All")
# Select cases for analysis
Sap.Scripts.Analyze.AddCases(Casename = ['DEAD', 'MODAL','yourCase1', 'yourCase2'])
# Delete Results
Sap.Scripts.Analyze.DeleteResults("All")
# Run analysis
Sap.Scripts.Analyze.RunAll()


# post process
# open excel
filename='F:\python\Sap2000\Models\Test.xlsx'
wb = openpyxl.load_workbook(filename)
# choose a target sheet
print("SheetNames are: ",wb.sheetnames)

# Open 
TargetSheet = wb.sheetnames[0]  # you can also put sheetnames here
Targetid = wb.sheetnames.index(TargetSheet)
ws=wb.worksheets[Targetid]
# change a name if you want
# ws.title = "yoursheetname"

# get reactions under deadload 
# select combo for output
Sap.Scripts.SelectCombo_Case(["DEAD"])

# get Frame force result by group name
Name,EleAbsForce,__,__ = Sap.Scripts.GetResults.ElementJointForce_by_Group("PierBottom")
# write in excel (here we need F3 --> [2]),"D22" is the top left corner of the matrix
Sap.Scripts.writecell(ws,EleAbsForce[:,[2]],"D22")

# get reactions under earthquake time history
# select combo for output
Sap.Scripts.SelectCombo_Case(["E2YEarthquake"])

# get Frame force result by group name
Name,EleAbsForce,EleMaxForce,EleMinForce = Sap.Scripts.GetResults.ElementJointForce_by_Group("PierBottom")
# write in excel (here we need F3,F1,M2 --> [2,0,4]),,"D30" is the top left corner of the matrix
Sap.Scripts.writecell(ws,EleAbsForce[:,[2,0,4]],"D30")


wb.save(filename)



# Save your file with a Filename(default: your ModelPath)
Sap.File.Save()

# Don't forget to close the program
Sap.closeSap()


